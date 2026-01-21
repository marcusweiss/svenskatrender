import argparse
import json
import re
import unicodedata
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


NS = {
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}

CHART_TYPE_TAGS = [
    ("line", ".//c:lineChart"),
    ("bar", ".//c:barChart"),
    ("area", ".//c:areaChart"),
    ("scatter", ".//c:scatterChart"),
    ("pie", ".//c:pieChart"),
    ("doughnut", ".//c:doughnutChart"),
    ("bubble", ".//c:bubbleChart"),
]


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value.strip().lower())
    value = "".join(ch for ch in value if not unicodedata.category(ch).startswith("M"))
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "section"


def is_section_name(name: str) -> bool:
    """Identify section headers - these are typically longer all-caps phrases.
    Short acronyms like 'EU', 'EURO', 'NATO', 'SR' should not be treated as sections."""
    stripped = name.replace(" ", "")
    # Must be all uppercase AND at least 8 characters (to exclude short acronyms)
    # Known section headers: SAMHÄLLSTRENDER, MEDIETRENDER, POLITISKA TRENDER, etc.
    return bool(stripped) and stripped == stripped.upper() and len(stripped) >= 8


def normalize_ref(ref: str) -> Tuple[str, str]:
    sheet, cells = ref.split("!")
    sheet = sheet.strip("'")
    return sheet, cells


def read_range(ws, cell_range: str) -> List:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    values = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        values.append([cell.value for cell in row])
    if min_col == max_col:
        return [row[0] for row in values]
    return values


def extract_text(element: Optional[ET.Element]) -> Optional[str]:
    if element is None:
        return None
    texts = []
    for node in element.findall(".//a:t", NS):
        if node.text:
            texts.append(node.text)
    if texts:
        return "".join(texts).strip()
    cache = element.find(".//c:strCache", NS)
    if cache is None:
        return None
    pts = []
    for pt in cache.findall(".//c:pt", NS):
        if pt.find("c:v", NS) is not None and pt.find("c:v", NS).text:
            pts.append(pt.find("c:v", NS).text)
    return " ".join(pts).strip() if pts else None


def extract_ref_text(node: ET.Element, xpath: str) -> Optional[str]:
    for ref_type in ("c:numRef", "c:strRef", "c:multiLvlStrRef"):
        ref = node.find(f"{xpath}/{ref_type}/c:f", NS)
        if ref is not None and ref.text:
            return ref.text
    return None


def fetch_range(workbook, ref: str) -> Tuple[List, str]:
    sheet_name, cells = normalize_ref(ref)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' is not present in workbook")
    ws = workbook[sheet_name]
    return read_range(ws, cells), sheet_name


def detect_chart(root: ET.Element) -> Tuple[Optional[str], Optional[ET.Element]]:
    for chart_type, xpath in CHART_TYPE_TAGS:
        node = root.find(xpath, NS)
        if node is not None:
            return chart_type, node
    return None, None


def parse_chart(xml_bytes: bytes, workbook) -> Optional[Dict]:
    root = ET.fromstring(xml_bytes)
    chart_type, chart_node = detect_chart(root)
    if chart_node is None:
        return None
    title = extract_text(root.find(".//c:title", NS))
    series_data = []
    categories = None
    sheet_name = None

    for ser in chart_node.findall("c:ser", NS):
        label = extract_text(ser.find("c:tx", NS)) or ""
        cat_ref = extract_ref_text(ser, "c:cat")
        if cat_ref and categories is None:
            categories, sheet_name = fetch_range(workbook, cat_ref)
        val_ref = extract_ref_text(ser, "c:val")
        if not val_ref:
            continue
        values, val_sheet = fetch_range(workbook, val_ref)
        sheet_name = sheet_name or val_sheet
        series_data.append({"name": label, "values": values})

    if not series_data:
        return None

    if categories is None:
        cat_ref = extract_ref_text(chart_node, "c:cat")
        if cat_ref:
            categories, sheet_name = fetch_range(workbook, cat_ref)

    # Convert title from ALL CAPS to sentence case if needed
    display_title = title
    if title:
        # Check if it's all uppercase or mostly uppercase
        if title.isupper() or (len(title) > 3 and title.upper() == title):
            # Special case for "PARTILEDARPOPULARITET"
            if "partiledarpopularitet" in title.lower():
                # Extract the part before parentheses if any
                if "(" in title:
                    parts = title.split("(")
                    main_part = parts[0].strip()
                    rest = "(" + parts[1] if len(parts) > 1 else ""
                    display_title = main_part.capitalize() + rest
                else:
                    display_title = title.capitalize()
            else:
                display_title = title.capitalize()
    
    return {
        "title": display_title,
        "sheet": sheet_name,
        "type": chart_type,
        "categories": categories,
        "series": series_data,
    }


def extract_table_data(ws) -> Optional[List[List[Optional[str]]]]:
    """Extract table data from worksheet, skipping metadata rows.
    Looks for the actual data table, avoiding rows with 'Rubrik', 'Fråga', 'Kommentar', 'Typ', etc."""
    # Metadata labels to skip
    metadata_labels = {"rubrik", "fråga", "kommentar", "kommentar2", "bortkodning", "typ", "enhet", "källa", "underrubrik", "frågeformulering"}
    
    table_data = []
    max_rows_to_check = 200
    max_cols = 50  # Increased for tables with many columns
    
    # Find the actual data table - look for a row that looks like a header (contains "År" or year numbers)
    start_row = None
    for row_idx in range(1, min(max_rows_to_check + 1, ws.max_row + 1)):
        row = ws[row_idx]
        row_values = []
        for col_idx in range(1, min(max_cols + 1, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                cell_str = str(cell.value).strip().lower()
                # Skip rows that contain metadata labels
                if cell_str in metadata_labels:
                    break
                row_values.append(str(cell.value).strip())
            else:
                row_values.append("")
        
        # Check if this looks like a table header (contains "År" or starts with a year-like number)
        if row_values:
            first_cell = row_values[0].lower() if row_values[0] else ""
            # This is likely the header row if it contains "år" or starts with a 4-digit year
            if "år" in first_cell or (first_cell.isdigit() and len(first_cell) == 4 and 1980 <= int(first_cell) <= 2100):
                start_row = row_idx
                break
    
    if start_row is None:
        return None
    
    # Extract table starting from the header row
    for row_idx in range(start_row, min(start_row + 100, ws.max_row + 1)):
        row = ws[row_idx]
        row_data = []
        is_metadata_row = False
        
        for col_idx in range(1, min(max_cols + 1, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                cell_str = str(cell.value).strip().lower()
                # Skip this row if it contains metadata labels
                if cell_str in metadata_labels:
                    is_metadata_row = True
                    break
                # Also skip rows that start with "kommentar" or "bortkodning" (case-insensitive)
                if cell_str.startswith("kommentar") or cell_str.startswith("bortkodning"):
                    is_metadata_row = True
                    break
                
                # Convert to string, handling dates and numbers
                if isinstance(cell.value, (int, float)):
                    row_data.append(str(cell.value))
                elif hasattr(cell.value, 'strftime'):  # datetime
                    row_data.append(cell.value.strftime('%Y'))
                else:
                    row_data.append(str(cell.value).strip())
            else:
                row_data.append(None)
        
        # Skip metadata rows
        if is_metadata_row:
            continue
        
        # Only add row if it has at least one non-None value
        if any(row_data):
            # Trim trailing None values
            while row_data and row_data[-1] is None:
                row_data.pop()
            if row_data:
                table_data.append(row_data)
    
    return table_data if len(table_data) > 1 else None  # Need at least header + 1 row


def extract_metadata(ws) -> Dict[str, Optional[str]]:
    """Extract 'Rubrik', 'Underrubrik', 'Fråga', 'Kommentar', 'Typ', and 'Källa' from worksheet."""
    metadata = {"rubrik": None, "underrubrik": None, "fraga": None, "kommentar": None, "typ": None, "kalla": None}
    
    # Search more thoroughly - check first 300 rows and 30 columns
    for row_idx in range(1, min(301, ws.max_row + 1)):
        for col_idx in range(1, min(31, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, str):
                cell_value = str(cell.value).strip()
                
                # Check for "Fråga" - exact match (case-insensitive)
                if cell_value.lower() == "fråga" and not metadata["fraga"]:
                    # Try next cell in same row (most common pattern: "Fråga" | "Question text")
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            # Skip if it's just "Kommentar" or empty
                            if next_value and next_value.lower() not in ["kommentar", "fråga", ""]:
                                metadata["fraga"] = next_value
                                break
                    # Also try cell below (sometimes label is above value)
                    if not metadata["fraga"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value and below_value.lower() not in ["kommentar", "fråga", ""]:
                                metadata["fraga"] = below_value
                
                # Check for "Kommentar" - exact match (case-insensitive)
                if cell_value.lower() == "kommentar" and not metadata["kommentar"]:
                    # Try next cell in same row (most common pattern: "Kommentar" | "Comment text")
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            # Skip if it's just "Fråga" or empty
                            if next_value and next_value.lower() not in ["fråga", "kommentar", ""]:
                                metadata["kommentar"] = next_value
                                break
                    # Also try cell below (sometimes label is above value)
                    if not metadata["kommentar"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value and below_value.lower() not in ["fråga", "kommentar", ""]:
                                metadata["kommentar"] = below_value
                
                # Check for "Typ" - exact match (case-insensitive)
                if cell_value.lower() == "typ" and not metadata["typ"]:
                    # Try next cell in same row (most common pattern: "Typ" | "Tabell" or "Diagram")
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            if next_value:
                                metadata["typ"] = next_value
                                break
                    # Also try cell below
                    if not metadata["typ"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value:
                                metadata["typ"] = below_value
                
                # Check for "Rubrik" - exact match (case-insensitive)
                if cell_value.lower() == "rubrik" and not metadata["rubrik"]:
                    # Try next cell in same row (most common pattern: "Rubrik" | "Heading text")
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            if next_value and next_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", ""]:
                                metadata["rubrik"] = next_value
                                break
                    # Also try cell below
                    if not metadata["rubrik"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value and below_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", ""]:
                                metadata["rubrik"] = below_value
                
                # Check for "Underrubrik" - exact match (case-insensitive)
                if cell_value.lower() == "underrubrik" and not metadata["underrubrik"]:
                    # Try next cell in same row
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            if next_value and next_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", "underrubrik", ""]:
                                metadata["underrubrik"] = next_value
                                break
                    # Also try cell below
                    if not metadata["underrubrik"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value and below_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", "underrubrik", ""]:
                                metadata["underrubrik"] = below_value
                
                # Check for "Källa" - exact match (case-insensitive)
                if cell_value.lower() == "källa" and not metadata["kalla"]:
                    # Try next cell in same row (most common pattern: "Källa" | "Source text")
                    for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        next_cell = ws.cell(row_idx, next_col)
                        if next_cell.value:
                            next_value = str(next_cell.value).strip()
                            # Skip if it's just another metadata label or empty
                            if next_value and next_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", "underrubrik", "källa", ""]:
                                metadata["kalla"] = next_value
                                break
                    # Also try cell below (sometimes label is above value)
                    if not metadata["kalla"] and row_idx < ws.max_row:
                        below_cell = ws.cell(row_idx + 1, col_idx)
                        if below_cell.value:
                            below_value = str(below_cell.value).strip()
                            if below_value and below_value.lower() not in ["fråga", "kommentar", "typ", "rubrik", "underrubrik", "källa", ""]:
                                metadata["kalla"] = below_value
    
    return metadata


def build_sections(workbook, charts: Dict[str, List[Dict]]) -> List[Dict]:
    sections = []
    current_section = None
    processed_sheets = []

    for sheet_name in workbook.sheetnames:
        if is_section_name(sheet_name):
            # Expand abbreviations in section titles
            section_title = sheet_name
            # Handle various possible spellings/cases
            if sheet_name.upper().replace(" ", "") == "POLSAKFRÅGOR" or sheet_name.upper() == "POL SAKFRÅGOR":
                section_title = "POLITISKA SAKFRÅGOR"
            
            section = {
                "title": section_title,
                "slug": slugify(sheet_name),
                "indicators": [],
            }
            sections.append(section)
            current_section = section
            continue

        chart_list = charts.get(sheet_name, [])
        # Include sheet even if it has no charts (might have metadata or table data)
        ws = workbook[sheet_name]
        metadata = extract_metadata(ws)
        
        # Extract table data only if Typ says "Tabell"
        table_data = None
        if metadata.get("typ") and metadata["typ"].lower().strip() == "tabell":
            table_data = extract_table_data(ws)
        
        # Check if we need to reverse the order (latest year first)
        # For "Medborgarnas viktigaste samhällsproblem", "Oro: Samtliga områden", and "Partisympati"
        needs_reverse = False
        raw_rubrik = metadata.get("rubrik", "").lower() if metadata.get("rubrik") else ""
        raw_underrubrik_check = metadata.get("underrubrik", "").lower() if metadata.get("underrubrik") else ""
        raw_sheet_name = sheet_name.lower() if sheet_name else ""
        
        if "medborgarnas viktigaste samhällsproblem" in raw_rubrik:
            needs_reverse = True
        elif "vad svenskar oroar sig" in raw_rubrik and "samtliga områden" in raw_underrubrik_check:
            # Only reverse for "Samtliga områden", NOT for "Ekonomisk kris och Stor arbetslöshet"
            needs_reverse = True
        elif "partisympati" in raw_rubrik or "partisymp" in raw_sheet_name:
            needs_reverse = True
        
        # Reverse chart data if needed
        if needs_reverse and chart_list:
            for chart in chart_list:
                if chart.get("categories"):
                    chart["categories"] = list(reversed(chart["categories"]))
                if chart.get("series"):
                    for serie in chart["series"]:
                        if serie.get("values"):
                            serie["values"] = list(reversed(serie["values"]))
        
        # Reverse table data if needed (keep header row)
        if needs_reverse and table_data and len(table_data) > 1:
            header = table_data[0]
            data_rows = table_data[1:]
            table_data = [header] + list(reversed(data_rows))

        if current_section is None:
            current_section = {
                "title": "Report",
                "slug": "report",
                "indicators": [],
            }
            sections.append(current_section)

        # Always include the sheet, even if it has no charts or metadata
        # This ensures all sheets from Excel are included
        # Use Rubrik for title if available, otherwise use sheet name
        # Convert from ALL CAPS to sentence case (only first letter capitalized, rest lowercase)
        raw_title = metadata["rubrik"] if metadata["rubrik"] else sheet_name
        
        # Normalize the title - handle both rubrik and sheet_name cases
        # Also handle special case where title might be "PARTILEDARPOPULARITET" or similar
        if raw_title:
            # Check if it's all uppercase (handles both single words and phrases)
            # Also check if it's mostly uppercase (like "PARTILEDARPOPULARITET")
            title_upper = raw_title.upper()
            title_lower = raw_title.lower()
            is_mostly_upper = raw_title.isupper() or (len(raw_title) > 3 and raw_title == title_upper and raw_title != title_lower)
            
            if is_mostly_upper:
                # Convert to sentence case: first letter uppercase, rest lowercase
                # But preserve proper nouns like "Sveriges" - handle common Swedish proper nouns
                words = raw_title.split()
                if len(words) > 0:
                    # First word: capitalize first letter, lowercase rest
                    words[0] = words[0].capitalize()
                    # Rest of words: lowercase, except for known proper nouns
                    proper_nouns = {"Sveriges", "Sverige", "Svenska", "Svenskar", "Svensk", "EU", "NATO"}
                    for i in range(1, len(words)):
                        word_lower = words[i].lower()
                        if word_lower.capitalize() in proper_nouns or words[i] in proper_nouns:
                            words[i] = words[i].capitalize()
                        else:
                            words[i] = words[i].lower()
                    display_title = " ".join(words)
                else:
                    # Single word - just capitalize first letter
                    display_title = raw_title.capitalize()
            else:
                display_title = raw_title
        else:
            # Handle sheet name if no rubrik
            if sheet_name:
                # Check if sheet name is uppercase
                if sheet_name.isupper() or (len(sheet_name) > 3 and sheet_name.upper() == sheet_name):
                    display_title = sheet_name.capitalize()
                else:
                    display_title = sheet_name
            else:
                display_title = sheet_name
        
        # Convert underrubrik to sentence case if needed
        raw_underrubrik = metadata["underrubrik"]
        display_underrubrik = None
        if raw_underrubrik:
            if raw_underrubrik.isupper():
                display_underrubrik = raw_underrubrik.capitalize()
            else:
                display_underrubrik = raw_underrubrik
            # Special case: Change "Arbetslöshet och ekonomisk kris" to "Samtliga områden"
            # But for the first "Oro" indicator (sheet "Oro 1"), change to "Ekonomisk kris och Stor arbetslöshet"
            if display_underrubrik == "Arbetslöshet och ekonomisk kris":
                if sheet_name.lower() == "oro 1":
                    display_underrubrik = "Ekonomisk kris och Stor arbetslöshet"
                else:
                    display_underrubrik = "Samtliga områden"
        
        # Special case: If title is "Förtroende För Samhällsinstitutioner" and has underrubrik,
        # change title to just "Förtroende"
        if display_title and "förtroende" in display_title.lower() and "samhällsinstitutioner" in display_title.lower():
            if display_underrubrik:
                display_title = "Förtroende"
        
        # Special case: Force "PARTILEDARPOPULARITET" to "Partiledarpopularitet"
        if display_title and "partiledarpopularitet" in display_title.lower():
            display_title = "Partiledarpopularitet"
        
        # Include underrubrik in the indicator data (frontend will append it to title when needed)
        indicator = {
            "title": display_title,
            "slug": slugify(sheet_name),  # Keep slug based on sheet name for consistency
            "sheet": sheet_name,
            "charts": chart_list,
            "table": table_data,  # Include table data if Typ == "Tabell"
            "typ": metadata["typ"],  # Include type field
            "rubrik": metadata["rubrik"],  # Include rubrik field
            "underrubrik": display_underrubrik,  # Include underrubrik field (converted to title case)
            "fraga": metadata["fraga"],
            "kommentar": metadata["kommentar"],
            "kalla": metadata["kalla"],  # Include källa field
        }
        current_section["indicators"].append(indicator)
        processed_sheets.append(sheet_name)

    # Return all sections, but only if they have at least one indicator
    # This ensures we don't create empty sections, but all sheets are included
    return [section for section in sections if section["indicators"]]


def export(workbook_path: Path, output_path: Path) -> Dict:
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)
    chart_map: Dict[str, List[Dict]] = defaultdict(list)

    with zipfile.ZipFile(workbook_path, "r") as archive:
        chart_files = sorted(
            name for name in archive.namelist() if name.startswith("xl/charts/chart") and name.endswith(".xml")
        )
        for chart_file in chart_files:
            chart_obj = parse_chart(archive.read(chart_file), wb)
            if not chart_obj or not chart_obj.get("sheet"):
                continue
            chart_obj["id"] = Path(chart_file).stem
            chart_obj["source"] = chart_file
            chart_map[chart_obj["sheet"]].append(chart_obj)

    sections = build_sections(wb, chart_map)
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_workbook": workbook_path.name,
        "section_count": len(sections),
        "sections": sections,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def main():
    parser = argparse.ArgumentParser(description="Export Svenska Trender charts to JSON.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("3. Svenska trender 1986-2024.xlsm"),
        help="Path to the macro-enabled Excel workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/report-data.json"),
        help="Destination JSON path.",
    )
    args = parser.parse_args()
    payload = export(args.workbook, args.output)
    
    # Debug: list all indicators
    total_indicators = sum(len(section["indicators"]) for section in payload["sections"])
    print(f"Wrote {args.output} with {payload['section_count']} sections and {total_indicators} indicators.")
    
    # List all sheet names processed
    all_sheets = []
    for section in payload["sections"]:
        for indicator in section["indicators"]:
            all_sheets.append(indicator["sheet"])
    
    # Check for missing sheets
    wb = load_workbook(args.workbook, data_only=True, keep_vba=True)
    all_excel_sheets = set(wb.sheetnames)
    processed_sheets_set = set(all_sheets)
    missing = all_excel_sheets - processed_sheets_set
    if missing:
        missing_list = ', '.join(sorted(missing))
        print(f"WARNING: Missing sheets ({len(missing)}): {missing_list}")
    else:
        print(f"SUCCESS: All {len(all_excel_sheets)} sheets processed successfully")
    print(f"Last 10 processed: {', '.join(all_sheets[-10:])}")


if __name__ == "__main__":
    main()

